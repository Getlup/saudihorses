# -*- coding: utf-8 -*-
"""
توليد ملف Excel (XLSX) لبيانات الحصان + التقاريره اليومية ضمن نطاق تاريخ محدد.
يستخدم openpyxl مباشرة (بدون pandas) لتفادي أي اعتماديات إضافية غير ضرورية.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="0D1C16", end_color="0D1C16", fill_type="solid")
HEADER_FONT = Font(color="F1EBDD", bold=True)
RTL_ALIGN = Alignment(horizontal="right", readingOrder=2, wrap_text=True, vertical="top")

STATUS_FIELDS = [
    ("appetite", "الشهية والتغذية"), ("water", "شرب الماء"), ("droppings", "الروث والتبول"),
    ("behavior", "الحالة والسلوك العام"), ("movement", "الحركة (عرج/إصابة)"),
]


def _style_header_row(ws, row_idx, n_cols):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = RTL_ALIGN


def build_horse_reports_xlsx(stable, horse, reports, range_label):
    """يبني ملف Excel من ورقتين: ملخص بيانات الحصان، وجدول التقارير اليومية ضمن النطاق المحدد."""
    wb = Workbook()
    wb.properties.creator = "Saudi Horses"

    # ---------------- ورقة 1: بيانات الحصان ----------------
    ws1 = wb.active
    ws1.title = "بيانات الحصان"
    ws1.sheet_view.rightToLeft = True

    profile_rows = [
        ("الاسم", horse.name),
        ("المربط", stable.name_ar),
        ("العمر", f"{horse.age} سنة" if horse.age is not None else "—"),
        ("الجنس", horse.gender or "—"),
        ("السلالة", horse.breed or "—"),
        ("اللون", horse.color or "—"),
        ("الأب", horse.sire_name or "—"),
        ("الأم", horse.dam_name or "—"),
        ("المالك", horse.owner.name if horse.owner else stable.name_ar),
        ("رقم الشريحة/الجواز", horse.chip_number or "—"),
        ("رقم الحظيرة", horse.stall_number or "—"),
        ("الحالة الصحية المهمة", horse.health_notes or "—"),
        ("الحساسية", horse.allergies or "—"),
        ("خطة التغذية", horse.feeding_plan or "—"),
        ("الطبيب البيطري", horse.vet_name or "—"),
        ("تواصل البيطري", horse.vet_contact or "—"),
        ("تنبيه مهم", horse.important_alert or "—"),
        ("نطاق التصدير", range_label),
    ]
    for i, (label, value) in enumerate(profile_rows, start=1):
        ws1.cell(row=i, column=1, value=label).alignment = RTL_ALIGN
        ws1.cell(row=i, column=1).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=value).alignment = RTL_ALIGN
    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 48

    # ---------------- ورقة 2: التقارير اليومية ----------------
    ws2 = wb.create_sheet("التقارير اليومية")
    ws2.sheet_view.rightToLeft = True

    headers = ["التاريخ"] + [label for _, label in STATUS_FIELDS] + \
        ["التدريب/النشاط", "الأدوية/العلاجات", "ملاحظة", "بواسطة"]
    for col, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=col, value=h)
    _style_header_row(ws2, 1, len(headers))

    for r_idx, report in enumerate(reports, start=2):
        col = 1
        ws2.cell(row=r_idx, column=col, value=report.report_date.strftime("%Y-%m-%d")).alignment = RTL_ALIGN
        col += 1
        for key, _ in STATUS_FIELDS:
            status = getattr(report, f"{key}_status")
            detail = getattr(report, f"{key}_detail")
            status_ar = "غير طبيعي" if status == "abnormal" else "طبيعي"
            value = f"{status_ar} — {detail}" if status == "abnormal" and detail else status_ar
            cell = ws2.cell(row=r_idx, column=col, value=value)
            cell.alignment = RTL_ALIGN
            if status == "abnormal":
                cell.font = Font(color="C9836E")
            col += 1
        ws2.cell(row=r_idx, column=col, value=report.training_activity or "—").alignment = RTL_ALIGN
        col += 1
        ws2.cell(row=r_idx, column=col, value=report.medication_given or "—").alignment = RTL_ALIGN
        col += 1
        ws2.cell(row=r_idx, column=col, value=report.note or "—").alignment = RTL_ALIGN
        col += 1
        ws2.cell(row=r_idx, column=col, value=report.staff.name if report.staff else "—").alignment = RTL_ALIGN

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
